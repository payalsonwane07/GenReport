from flask import Flask, render_template, request, jsonify, send_from_directory
from werkzeug.utils import secure_filename
import os
import json
from datetime import datetime
import traceback
import pandas as pd
import numpy as np

from config import config

def create_app(env='development'):
    """Create Flask application"""
    app = Flask(__name__)
    app.config.from_object(config[env])
    
    # Create required folders
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['REPORTS_FOLDER'], exist_ok=True)
    
    return app

app = create_app('development')

# =====================================================
# ROUTES
# =====================================================

@app.route('/')
def index():
    """Home page"""
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    """Dashboard page"""
    return render_template('dashboard.html')

@app.route('/preview')
def preview():
    """Preview page"""
    return render_template('preview.html')

# =====================================================
# API ENDPOINTS
# =====================================================

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Handle file upload and initial processing"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'File type not allowed. Use CSV, Excel, or JSON'}), 400
        
        # Save file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_")
        filename = timestamp + filename
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Load and analyze data
        try:
            if filepath.endswith('.csv'):
                df = pd.read_csv(filepath)
            elif filepath.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(filepath)
            elif filepath.endswith('.json'):
                df = pd.read_json(filepath)
            
            # Clean data
            df = df.drop_duplicates()
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            df[numeric_cols] = df[numeric_cols].fillna(df[numeric_cols].mean())
            
            # Calculate metrics
            total_rows = len(df)
            total_cols = len(df.columns)
            missing_values = int(df.isnull().sum().sum())
            data_quality = ((total_rows * total_cols - missing_values) / (total_rows * total_cols) * 100)
            
            return jsonify({
                'success': True,
                'file_id': filename,
                'filename': file.filename,
                'rows': total_rows,
                'columns': total_cols,
                'column_names': list(df.columns),
                'dtypes': df.dtypes.to_dict(),
                'missing_values': missing_values,
                'data_quality': round(data_quality, 2)
            }), 200
        
        except Exception as e:
            return jsonify({'error': f'Failed to process file: {str(e)}'}), 400
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/preview/<file_id>')
def preview_data(file_id):
    """Get data preview"""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file_id))
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        # Load data
        if filepath.endswith('.csv'):
            df = pd.read_csv(filepath)
        elif filepath.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(filepath)
        elif filepath.endswith('.json'):
            df = pd.read_json(filepath)
        
        # Get preview
        preview_data = df.head(10).to_dict('records')
        
        # Get statistics
        numeric_df = df.select_dtypes(include=[np.number])
        stats = numeric_df.describe().to_dict()
        
        return jsonify({
            'preview': preview_data,
            'total_rows': len(df),
            'stats': stats,
            'columns': list(df.columns)
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analyze/<file_id>')
def analyze_data(file_id):
    """Perform data analysis"""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file_id))
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        # Load data
        if filepath.endswith('.csv'):
            df = pd.read_csv(filepath)
        elif filepath.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(filepath)
        elif filepath.endswith('.json'):
            df = pd.read_json(filepath)
        
        # Numeric columns only
        numeric_df = df.select_dtypes(include=[np.number])
        
        # Correlations
        correlation_matrix = numeric_df.corr()
        high_correlations = {}
        
        for i in range(len(correlation_matrix.columns)):
            for j in range(i+1, len(correlation_matrix.columns)):
                corr_value = correlation_matrix.iloc[i, j]
                if abs(corr_value) >= 0.5:
                    col1 = correlation_matrix.columns[i]
                    col2 = correlation_matrix.columns[j]
                    high_correlations[f"{col1} ↔ {col2}"] = round(float(corr_value), 3)
        
        # Summary statistics
        summary = {
            'total_records': len(df),
            'total_columns': len(df.columns),
            'numeric_columns': len(numeric_df.columns),
            'categorical_columns': len(df.select_dtypes(include=['object']).columns),
            'missing_values': int(df.isnull().sum().sum()),
            'completeness': round((1 - df.isnull().sum().sum() / (len(df) * len(df.columns))) * 100, 2)
        }
        
        # Basic statistics for each numeric column
        column_stats = {}
        for col in numeric_df.columns[:5]:  # First 5 columns
            column_stats[col] = {
                'mean': round(float(numeric_df[col].mean()), 2),
                'median': round(float(numeric_df[col].median()), 2),
                'std': round(float(numeric_df[col].std()), 2),
                'min': round(float(numeric_df[col].min()), 2),
                'max': round(float(numeric_df[col].max()), 2)
            }
        
        return jsonify({
            'summary': summary,
            'correlations': high_correlations,
            'column_stats': column_stats,
            'success': True
        }), 200
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/generate-report', methods=['POST'])
def generate_report():
    """Generate PDF/Excel report"""
    try:
        data = request.json
        file_id = data.get('file_id')
        report_type = data.get('report_type', 'pdf')
        title = data.get('title', 'Data Analysis Report')
        
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file_id))
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        # Load data
        if filepath.endswith('.csv'):
            df = pd.read_csv(filepath)
        elif filepath.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(filepath)
        elif filepath.endswith('.json'):
            df = pd.read_json(filepath)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if report_type == 'pdf':
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer, PageBreak
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.lib import colors
                from reportlab.lib.enums import TA_CENTER
                
                report_filename = f"report_{timestamp}.pdf"
                report_path = os.path.join(app.config['REPORTS_FOLDER'], report_filename)
                
                doc = SimpleDocTemplate(report_path, pagesize=letter)
                story = []
                styles = getSampleStyleSheet()
                
                # Title
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=24,
                    textColor=colors.HexColor('#1f4788'),
                    spaceAfter=30,
                    alignment=TA_CENTER
                )
                
                story.append(Paragraph(title, title_style))
                story.append(Paragraph(f"<i>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>", styles['Normal']))
                story.append(Spacer(1, 0.3*inch))
                
                # Executive Summary
                story.append(Paragraph("Executive Summary", styles['Heading2']))
                summary_text = f"This report analyzes {len(df)} records with {len(df.columns)} columns."
                story.append(Paragraph(summary_text, styles['Normal']))
                story.append(Spacer(1, 0.3*inch))
                
                # Data Summary
                story.append(PageBreak())
                story.append(Paragraph("Data Summary", styles['Heading2']))
                
                numeric_df = df.select_dtypes(include=[np.number])
                summary_text = f"""
                <b>Total Records:</b> {len(df)}<br/>
                <b>Total Columns:</b> {len(df.columns)}<br/>
                <b>Numeric Columns:</b> {len(numeric_df.columns)}<br/>
                <b>Missing Values:</b> {int(df.isnull().sum().sum())}<br/>
                """
                story.append(Paragraph(summary_text, styles['Normal']))
                story.append(Spacer(1, 0.3*inch))
                
                # Data Table
                story.append(Paragraph("Data Preview (First 10 Rows)", styles['Heading2']))
                display_data = df.head(10)
                table_data = [list(display_data.columns)] + display_data.values.tolist()
                table = Table(table_data)
                table.setStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ])
                story.append(table)
                
                # Build PDF
                doc.build(story)
                
                return jsonify({
                    'success': True,
                    'report_filename': report_filename,
                    'download_url': f'/api/download/{report_filename}',
                    'message': f'PDF report generated successfully!'
                }), 200
            
            except ImportError:
                return jsonify({'error': 'PDF library not installed. Run: pip install reportlab'}), 500
        
        else:  # Excel
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                report_filename = f"report_{timestamp}.xlsx"
                report_path = os.path.join(app.config['REPORTS_FOLDER'], report_filename)
                
                wb = Workbook()
                wb.remove(wb.active)
                
                # Data sheet
                ws = wb.create_sheet('Data')
                for r_idx, row in enumerate(df.values, 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Add headers with formatting
                for c_idx, col in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=c_idx)
                    cell.value = col
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
                
                # Summary sheet
                ws_summary = wb.create_sheet('Summary')
                ws_summary['A1'] = 'Metric'
                ws_summary['B1'] = 'Value'
                
                ws_summary['A1'].font = Font(bold=True)
                ws_summary['B1'].font = Font(bold=True)
                
                ws_summary['A2'] = 'Total Records'
                ws_summary['B2'] = len(df)
                ws_summary['A3'] = 'Total Columns'
                ws_summary['B3'] = len(df.columns)
                ws_summary['A4'] = 'Missing Values'
                ws_summary['B4'] = int(df.isnull().sum().sum())
                
                wb.save(report_path)
                
                return jsonify({
                    'success': True,
                    'report_filename': report_filename,
                    'download_url': f'/api/download/{report_filename}',
                    'message': f'Excel report generated successfully!'
                }), 200
            
            except ImportError:
                return jsonify({'error': 'Excel library not installed. Run: pip install openpyxl'}), 500
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/download/<filename>')
def download_report(filename):
    """Download generated report"""
    try:
        filename = secure_filename(filename)
        return send_from_directory(app.config['REPORTS_FOLDER'], filename, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reports')
def list_reports():
    """List all generated reports"""
    try:
        reports = []
        if os.path.exists(app.config['REPORTS_FOLDER']):
            for file in os.listdir(app.config['REPORTS_FOLDER']):
                file_path = os.path.join(app.config['REPORTS_FOLDER'], file)
                if os.path.isfile(file_path):
                    reports.append({
                        'filename': file,
                        'size': os.path.getsize(file_path),
                        'created': datetime.fromtimestamp(os.path.getctime(file_path)).isoformat(),
                        'download_url': f'/api/download/{file}'
                    })
        
        return jsonify({'reports': sorted(reports, key=lambda x: x['created'], reverse=True)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.errorhandler(404)
def not_found(error):
    """Handle 404 errors"""
    return jsonify({'error': 'Not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors"""
    return jsonify({'error': 'Internal server error'}), 500

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

if __name__ == '__main__':
    print("=" * 60)
    print("🚀 Automated Report Generator")
    print("=" * 60)
    print()
    print("📊 Starting server...")
    print()
    print("🌐 Visit: http://localhost:5000")
    print()
    print("Press Ctrl+C to stop the server")
    print("=" * 60)
    print()
    app.run(debug=True, host='0.0.0.0', port=5000)
